"""
Flask Research Portal Application
Earnings Call / Management Commentary Analysis Tool
Frontend handles all file processing - backend only analyzes text
"""

from flask import Flask, request, jsonify, send_from_directory, make_response, send_file
from flask_cors import CORS
import os
import uuid
from dotenv import load_dotenv
from datetime import datetime
import io
from typing import Dict, Any

from utils.text_processor import chunk_text, needs_chunking
from utils.gemini_analyzer import analyze_document, analyze_financial_document
from utils.validator import validate_analysis_output, sanitize_output
from utils.finance_extractor import (
    extract_periods_from_text,
    extract_currency_and_unit,
    create_extraction_prompt,
    parse_extraction_result,
    calculate_derived_metrics,
    generate_excel_data
)

# Load environment variables
load_dotenv()

# Get GitHub token from environment
GITHUB_TOKEN = os.getenv('GITHUB_TOKEN')
if not GITHUB_TOKEN:
    raise ValueError("GITHUB_TOKEN not found in environment variables")

# Ensure it's a string (Render sometimes passes as different types)
GITHUB_TOKEN = str(GITHUB_TOKEN).strip()

app = Flask(__name__, static_folder='static')
CORS(app)  # Enable CORS for all routes

# Configuration
MAX_FILE_SIZE_MB = int(os.getenv('MAX_FILE_SIZE_MB', 20))
MAX_TEXT_LENGTH = MAX_FILE_SIZE_MB * 1024 * 1024  # Max text length in characters

# In-memory storage for document texts
documents = {}


@app.route('/')
def index():
    """Serve the main HTML page."""
    return send_from_directory('static', 'index.html')


@app.route('/upload', methods=['POST'])
def upload():
    """
    Receive extracted text from frontend.
    Frontend handles all file processing (PDF extraction, OCR, text reading).
    
    Request: JSON with 'text' and 'filename'
    Response: document_id and text_length
    """
    try:
        data = request.get_json()
        
        if not data or 'text' not in data:
            return jsonify({"error": "No text provided. Frontend must extract text from files."}), 400
        
        text = data['text']
        filename = data.get('filename', 'document.txt')
        
        # Validate text
        if not text or not text.strip():
            return jsonify({"error": "Empty text received. Please ensure file has content."}), 400
        
        if len(text) > MAX_TEXT_LENGTH:
            return jsonify({
                "error": f"Text too large. Maximum {MAX_FILE_SIZE_MB}MB allowed."
            }), 413
        
        # Generate unique document ID
        document_id = str(uuid.uuid4())
        
        # Store text in memory
        documents[document_id] = {
            'text': text,
            'filename': filename
        }
        
        return jsonify({
            "document_id": document_id,
            "filename": filename,
            "text_length": len(text),
            "status": "ready",
            "message": "Text received successfully"
        }), 200
        
    except Exception as e:
        return jsonify({"error": f"Upload failed: {str(e)}"}), 500


@app.route('/analyze', methods=['POST'])
def analyze():
    """
    Analyze uploaded document using GitHub Models AI.
    
    Request: JSON with document_id
    Response: Structured analysis result
    """
    # Check API key
    if not GITHUB_TOKEN:
        return jsonify({
            "error": "GITHUB_TOKEN not configured. Please set it in .env file"
        }), 500
    
    # Get document_id from request
    data = request.get_json()
    if not data or 'document_id' not in data:
        return jsonify({"error": "document_id required in request body"}), 400
    
    document_id = data['document_id']
    
    # Check if document exists
    if document_id not in documents:
        return jsonify({"error": "Document not found. Please upload first."}), 404
    
    try:
        # Get document text
        doc = documents[document_id]
        text = doc['text']
        
        # Check if chunking is needed (gpt-4o has 8000 token limit)
        chunks = None
        if needs_chunking(text, max_tokens=6000):  # Leave margin for prompt overhead
            chunks = chunk_text(text, max_tokens=6000)
        
        # Analyze using GitHub Models
        try:
            result = analyze_document(GITHUB_TOKEN, text, chunks)
        except Exception as api_error:
            return jsonify({
                "error": f"AI API error: {str(api_error)}",
                "error_type": type(api_error).__name__
            }), 500
        
        # Sanitize output
        try:
            result = sanitize_output(result)
        except Exception as sanitize_error:
            return jsonify({
                "error": f"Output sanitization error: {str(sanitize_error)}",
                "error_type": type(sanitize_error).__name__,
                "raw_result": str(result)[:500]  # First 500 chars for debugging
            }), 500
        
        # Validate output
        is_valid, error_msg = validate_analysis_output(result)
        if not is_valid:
            return jsonify({
                "error": f"Analysis output validation failed: {error_msg}",
                "raw_output": result
            }), 500
        
        # Remove from documents dictionary after successful analysis
        documents.pop(document_id, None)
        
        return jsonify({
            "document_id": document_id,
            "filename": doc['filename'],
            "analysis": result,
            "status": "completed"
        }), 200
        
    except Exception as e:
        return jsonify({
            "error": f"Analysis failed: {str(e)}",
            "error_type": type(e).__name__
        }), 500


@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint."""
    return jsonify({"status": "healthy", "service": "research-portal"}), 200


@app.route('/financial-extract', methods=['POST'])
def financial_extract():
    """
    Enhanced financial extraction with multi-period support and Excel output.
    
    Request: JSON with document_id
    Response: Excel file with Income Statement and Metadata sheets
    """
    if not GITHUB_TOKEN:
        return jsonify({"error": "GITHUB_TOKEN not configured"}), 500
    
    data = request.get_json()
    if not data or 'document_id' not in data:
        return jsonify({"error": "document_id required"}), 400
    
    document_id = data['document_id']
    
    if document_id not in documents:
        return jsonify({"error": "Document not found"}), 404
    
    try:
        doc = documents[document_id]
        text = doc['text']
        filename = doc['filename']
        
        # Step 1: Extract periods from document
        periods = extract_periods_from_text(text)
        if not periods:
            return jsonify({
                "error": "No financial periods found. Ensure document contains period identifiers like 'FY25', 'Q4 FY25', etc."
            }), 400
        
        # Step 2: Detect currency and unit
        currency, unit = extract_currency_and_unit(text)
        
        # Step 3: Create extraction prompt
        prompt = create_extraction_prompt(text, periods)
        
        # Step 4: Call LLM for extraction
        try:
            extraction_result = analyze_financial_document(GITHUB_TOKEN, text, prompt)
        except Exception as e:
            return jsonify({"error": f"LLM extraction failed: {str(e)}"}), 500
        
        # Step 5: Convert to line items dict
        line_items = {}
        for item_data in extraction_result.get('line_items', []):
            line_items[item_data['name']] = {
                "name": item_data['name'],
                "values": item_data['values'],
                "status": {period: 'extracted' for period in item_data['values'].keys()}
            }
        
        # Add Total Revenue if we have components
        if "Revenue from operations" in line_items and "Other income" in line_items:
            total_revenue = {"name": "Total Revenue", "values": {}, "status": {}}
            for period in periods:
                ops = line_items["Revenue from operations"]["values"].get(period)
                other = line_items["Other income"]["values"].get(period)
                if ops is not None and other is not None:
                    total_revenue["values"][period] = ops + other
                    total_revenue["status"][period] = "calculated"
            line_items["Total Revenue"] = total_revenue
        
        # Step 6: Calculate derived metrics
        calculated_items = calculate_derived_metrics(line_items)
        line_items.update(calculated_items)
        
        # Step 7: Generate Excel data
        metadata = {
            'source_document': filename,
            'extraction_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'currency': extraction_result.get('currency', currency),
            'unit': extraction_result.get('unit', unit),
            'periods': extraction_result.get('periods', periods)
        }
        
        excel_data = generate_excel_data(line_items, metadata)
        
        # Step 8: Create Excel file
        excel_file = create_excel_file(excel_data, metadata)
        
        # Clean up
        documents.pop(document_id, None)
        
        # Send file
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'financial_statement_{document_id}.xlsx'
        )
        
    except Exception as e:
        return jsonify({
            "error": f"Financial extraction failed: {str(e)}",
            "error_type": type(e).__name__
        }), 500


def create_excel_file(data: Dict[str, Any], metadata: Dict[str, Any]) -> io.BytesIO:
    """Create professional Excel file with formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Income Statement sheet
    ws_income = wb.create_sheet("Income Statement", 0)
    income_data = data["Income Statement"]
    
    for row_idx, row_data in enumerate(income_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_income.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Header formatting
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Section headers
            elif isinstance(value, str) and value.endswith(":"):
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Line item names
            elif col_idx == 1 and row_idx > 1:
                cell.alignment = Alignment(horizontal="left", indent=1)
            
            # Numeric values
            elif col_idx > 1 and col_idx <= len(income_data[0]) - 2:
                if isinstance(value, str) and value.replace(",", "").replace(".", "").replace("-", "").isdigit():
                    if "Gross Margin" in income_data[row_idx-1][0]:
                        cell.number_format = '0.00%'
                        try:
                            cell.value = float(value)
                        except:
                            pass
                    else:
                        cell.number_format = '#,##0.00'
                        try:
                            cell.value = float(value.replace(",", ""))
                        except:
                            pass
                cell.alignment = Alignment(horizontal="right")
            
            # Status column
            elif col_idx == len(income_data[0]) - 1:
                if "✓" in str(value):
                    cell.font = Font(color="008000")
                elif "⚡" in str(value):
                    cell.font = Font(color="FFA500")
                elif "✗" in str(value):
                    cell.font = Font(color="FF0000")
    
    # Set column widths
    ws_income.column_dimensions['A'].width = 35
    for col in range(2, len(income_data[0]) + 1):
        ws_income.column_dimensions[chr(64 + col)].width = 15
    
    ws_income.freeze_panes = 'A2'
    
    # Metadata sheet
    ws_meta = wb.create_sheet("Metadata", 1)
    metadata_data = data["Metadata"]
    
    for row_idx, row_data in enumerate(metadata_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_meta.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            if col_idx == 1:
                cell.font = Font(bold=True)
    
    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 50
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', 'True') == 'True'
    app.run(host='0.0.0.0', port=port, debug=debug)